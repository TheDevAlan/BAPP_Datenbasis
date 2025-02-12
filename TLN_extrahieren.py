import os
import re
from openpyxl import load_workbook

# Pfad zum Quellordner
source_folder = 'C:\\Users\\alang\\Downloads\\Erstellung_BAPP\\Teilnehmer\\'

# Durchlaufen aller Dateien im Quellordner
for file_name in os.listdir(source_folder):
    file_path = os.path.join(source_folder, file_name)
    if os.path.isfile(file_path):
        projektnummer, tln_nummer = None, None
        lesen_soll_phasen, lesen_ist_phasen = False, False
        phasen_info = {'SOLL': [], 'IST': []}

        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                # Suchen nach "Projektnummer:"
                if not projektnummer and "Projektnummer:" in line:
                    # Extraktion der 10-stelligen Zahl mit regulären Ausdrücken
                    match = re.search(r'Projektnummer:\s*(\d{10})', line)
                    if match:
                        projektnummer = match.group(1)
                        print(f"Gefundene Projektnummer: {projektnummer}")

                # Suchen nach "TLN-Nummer" in Verbindung mit Projektnummer
                if projektnummer and "TLN-Nummer" in line:
                    match_tln = re.search(r'TLN-Nummer\s*TLN-(\d{10})', line)
                    if match_tln:
                        tln_nummer = match_tln.group(1)
                        print(f"Gefundene TLN-Nummer in Verbindung mit Projektnummer {projektnummer}: TLN-{tln_nummer}")
                
                # Prüfen, ob die SOLL Phaseninformationen beginnen
                if 'geplante Phasen (SOLL)' in line:
                    lesen_soll_phasen = True
                    lesen_ist_phasen = False
                    continue

                # Prüfen, ob die IST Phaseninformationen beginnen
                if 'tatsächliche Phasen (IST)' in line:
                    lesen_ist_phasen = True
                    lesen_soll_phasen = False
                    continue
                
                # Phaseninformationen für SOLL und IST lesen
                if lesen_soll_phasen or lesen_ist_phasen:
                    if line.strip() == '':
                        lesen_soll_phasen = False
                        lesen_ist_phasen = False
                        continue

                    phase_match = re.match(r'(\d{4} /\s\d{2})\s([TBG]{1})\s*:\s*(\w+)', line)
                    if phase_match:
                        monat, phase_kurz, phase_lang = phase_match.groups()
                        phase_typ = 'SOLL' if lesen_soll_phasen else 'IST'
                        phasen_info[phase_typ].append({
                            'projektnummer': projektnummer,
                            'tln_nummer': tln_nummer,
                            'monat': monat,
                            'phase_kurz': phase_kurz,
                            'phase_lang': phase_lang
                        })

            # Kombinierte Phasendaten
            kombinierte_phasen = {}

            for phase_typ in ['SOLL', 'IST']:
                for phase in phasen_info[phase_typ]:
                    key = (phase['monat'], phase['projektnummer'], phase['tln_nummer'])
                    if key not in kombinierte_phasen:
                        kombinierte_phasen[key] = {'SOLL': None, 'IST': None}
                    kombinierte_phasen[key][phase_typ] = phase['phase_kurz']


            # Excel-Datei öffnen
            excel_file = load_workbook('C:/Users/alang/Downloads/Erstellung_BAPP/Prüfvermerk.xlsx')
            sheet = excel_file.active

            # Letzte Zeile ermitteln, um neue Daten hinzuzufügen
            last_row = sheet.max_row + 1

            # Hinzufügen der SOLL- und IST-Daten und Formeln in Spalte F, G und H
            for key, phase in kombinierte_phasen.items():
                monat, projektnummer, tln_nummer = key
                sheet.cell(row=last_row, column=1).value = projektnummer
                sheet.cell(row=last_row, column=2).value = f"TLN-{tln_nummer}"
                sheet.cell(row=last_row, column=3).value = monat
                sheet.cell(row=last_row, column=4).value = phase['SOLL']
                sheet.cell(row=last_row, column=5).value = phase['IST']

                # Formeln in Spalte F, G und H hinzufügen
                sheet.cell(row=last_row, column=6).value = f'=LEFT(C{last_row}, FIND(" / ", C{last_row})-1)'
                sheet.cell(row=last_row, column=7).value = f'=VALUE(RIGHT(C{last_row}, LEN(C{last_row}) - FIND(" / ", C{last_row}) - 2))'
                sheet.cell(row=last_row, column=8).value = f'=B{last_row} & "-" & F{last_row} & "-" & G{last_row}'

                last_row += 1

            excel_file.save('C:\\Users\\alang\\Downloads\\Erstellung_BAPP\\Prüfvermerk.xlsx')